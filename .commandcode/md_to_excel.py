#!/usr/bin/env python
"""Convert TestCases_VWO_Login_Dashboard.md into a formatted Excel workbook."""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\user\Roopa M\AITESTERBLUEPRINT4X\Chapter 01_LLM Basics\TestCases_VWO_Login_Dashboard.md"
OUT = r"C:\Users\user\Roopa M\AITESTERBLUEPRINT4X\Chapter 01_LLM Basics\TestCases_VWO_Login_Dashboard.xlsx"

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FONT = Font(bold=True, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_tables(md):
    """Yield (heading, headers, rows) for every markdown table in md."""
    lines = md.splitlines()
    tables = []
    current_heading = ""
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("###"):
            current_heading = line.lstrip("#").strip()
        if line.strip().startswith("|"):
            # find table block: header, separator, data rows
            block = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                block.append(lines[i].strip())
                i += 1
            if len(block) >= 2 and re.match(r"^\|[\s:|-]+\|$", block[1]):
                headers = [c.strip() for c in block[0].strip("|").split("|")]
                rows = []
                for r in block[2:]:
                    if r.strip().startswith("|"):
                        rows.append([c.strip() for c in r.strip("|").split("|")])
                tables.append((current_heading, headers, rows))
                continue
        i += 1
    return tables


def style_header(ws, row, headers, widths):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 22)
    ws.row_dimensions[row].height = 22


def style_data(ws, start_row, rows):
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = WRAP
            c.border = BORDER


def main():
    with open(SRC, encoding="utf-8") as f:
        md = f.read()

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Test Cases (section 3)
    tables = parse_tables(md)
    ws = wb.create_sheet("Test Cases")
    ws.cell(row=1, column=1, value="Test Cases — VWO Login Dashboard")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.row_dimensions[1].height = 26

    row = 3
    for heading, headers, rows in tables:
        if not heading.startswith("3."):
            continue
        ws.cell(row=row, column=1, value=heading)
        ws.cell(row=row, column=1).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        row += 1
        style_header(ws, row, headers, {1: 10, 2: 14, 3: 40, 4: 30, 5: 60, 6: 50})
        row += 1
        style_data(ws, row, rows)
        row += len(rows) + 1
    ws.freeze_panes = "A3"

    # Sheet 2: Verified Facts (section 1)
    ws2 = wb.create_sheet("Verified Facts")
    ws2.cell(row=1, column=1, value="1. Verified Facts (Extracted from PRD Only)")
    ws2.cell(row=1, column=1).font = TITLE_FONT
    ws2.cell(row=1, column=1).fill = TITLE_FILL
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    for heading, headers, rows in tables:
        if heading.startswith("1."):
            style_header(ws2, 3, headers, {1: 8, 2: 90, 3: 40})
            style_data(ws2, 4, rows)
            break
    ws2.freeze_panes = "A4"

    # Sheet 3: Missing / Unknown (section 2)
    ws3 = wb.create_sheet("Missing Info")
    ws3.cell(row=1, column=1, value="2. Missing / Unknown Information (Not Stated in PRD)")
    ws3.cell(row=1, column=1).font = TITLE_FONT
    ws3.cell(row=1, column=1).fill = TITLE_FILL
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    for heading, headers, rows in tables:
        if heading.startswith("2."):
            style_header(ws3, 3, headers, {1: 6, 2: 70, 3: 60})
            style_data(ws3, 4, rows)
            break
    ws3.freeze_panes = "A4"

    # Sheet 4: Traceability (section 4)
    ws4 = wb.create_sheet("Traceability")
    ws4.cell(row=1, column=1, value="4. Output Traceability Map")
    ws4.cell(row=1, column=1).font = TITLE_FONT
    ws4.cell(row=1, column=1).fill = TITLE_FILL
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    for heading, headers, rows in tables:
        if heading.startswith("4."):
            style_header(ws4, 3, headers, {1: 45, 2: 45})
            style_data(ws4, 4, rows)
            break
    ws4.freeze_panes = "A4"

    wb.save(OUT)
    print("Saved:", OUT)
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
